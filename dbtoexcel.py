import psycopg2
import pandas as pd
import openpyxl
from openpyxl.styles import Font


def export_to_excel(connection, query_string, headings, filepath):

    """
    Exports data from PostgreSQL to an Excel spreadsheet using psycopg2.

    Arguments:
    connection - an open psycopg2 (this function does not close the connection)
    query_string - SQL to get data
    headings - list of strings to use as column headings
    filepath - path and filename of the Excel file

    psycopg2 and file handling errors bubble up to calling code.
    """

    cursor = connection.cursor()
    cursor.execute(query_string)
    data = cursor.fetchall()
    cursor.close()

    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.row_dimensions[1].font = Font(bold = True)

    # Spreadsheet row and column indexes start at 1
    # so we use "start = 1" in enumerate so
    # we don't need to add 1 to the indexes.
    #Enumerate() method adds a counter to an iterable and returns it in a form of enumerating object
    for colno, heading in enumerate(headings, start = 1):
        sheet.cell(row = 1, column = colno).value = heading

    # This time we use "start = 2" to skip the heading row.
    for rowno, row in enumerate(data, start = 2):
        for colno, cell_value in enumerate(row, start = 1):
            sheet.cell(row = rowno, column = colno).value = cell_value

    wb.save(filepath)
    
export_to_excel(connection=psycopg2.connect(host='localhost',database='Demo',user='postgres',password='root',port=5432),query_string="select * from araiindia1",headings=['SR_N','CODE','TITLE','AttachedFile'],filepath="ExportDataFromDBtoExcel.xlsx")


